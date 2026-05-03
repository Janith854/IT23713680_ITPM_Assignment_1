from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl

ROOT_DIR = Path(__file__).resolve().parent
TESTS_DIR = ROOT_DIR

DEFAULT_EXCEL_CANDIDATES = [
    str(TESTS_DIR / "Assignment 1 - Test cases.xlsx"),
]

DEFAULT_SHEET_NAME = "Test cases"
DEFAULT_FRONTEND_URL = "https://www.pixelssuite.com/chat-translator"

DEFAULT_OUTPUT_WAIT_MS = 45000
DEFAULT_TEXTAREA_WAIT_MS = 30000

def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="ignore")
    except:
        pass

def _normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())

def _find_column(header, name):
    for i, col in enumerate(header):
        if _normalize_header(col) == _normalize_header(name):
            return i + 1
    return None

def _wait_for_output_textarea(page, timeout_ms):
    # Wait for at least one textarea to be present
    page.wait_for_selector("textarea", timeout=timeout_ms)
    # Ensure both input and output textareas are loaded
    page.wait_for_function(
        "() => document.querySelectorAll('textarea').length >= 2",
        timeout=timeout_ms,
    )

def _read_output_with_wait(page, output_box, timeout_ms):
    try:
        output_box.wait_for(state="attached", timeout=timeout_ms)
        prev = output_box.input_value()
        handle = output_box.element_handle()
        if handle:
            page.wait_for_function(
                "(el, prev) => el && el.value !== prev && el.value.trim().length > 0",
                arg=[handle, prev],
                timeout=timeout_ms,
            )
    except PlaywrightTimeoutError:
        pass
    return output_box.input_value().strip()

def run_test():
    _configure_stdout()

    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=DEFAULT_EXCEL_CANDIDATES[0])
    parser.add_argument("--url", default=DEFAULT_FRONTEND_URL)
    parser.add_argument("--output-wait-ms", type=int, default=DEFAULT_OUTPUT_WAIT_MS)
    parser.add_argument("--textarea-wait-ms", type=int, default=DEFAULT_TEXTAREA_WAIT_MS)
    parser.add_argument("--keep-open", action="store_true")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"Excel file not found: {args.excel}")
        return

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[DEFAULT_SHEET_NAME] if DEFAULT_SHEET_NAME in wb.sheetnames else wb.active

    header = [cell.value for cell in ws[1]]

    input_col = _find_column(header, "Input")
    expected_col = _find_column(header, "Expected output")
    actual_col = _find_column(header, "Actual output")
    status_col = _find_column(header, "Status")

    if not all([input_col, expected_col, actual_col, status_col]):
        print("Missing one or more required columns: Input, Expected output, Actual output, Status")
        return

    print("Total rows to process:", ws.max_row - 1)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=200)
        page = browser.new_page()
        page.goto(args.url)
        page.wait_for_load_state("networkidle")
        
        # 1. Initialize selectors
        _wait_for_output_textarea(page, args.textarea_wait_ms)
        translate_btn = page.get_by_role("button", name=re.compile("Transliterate", re.I))

        input_box = page.locator("textarea").first
        output_box = page.locator("textarea").nth(1)

        for i in range(2, ws.max_row + 1):
            input_text = ws.cell(i, input_col).value
            if not input_text:
                continue

            input_text = str(input_text).strip()
            expected = ws.cell(i, expected_col).value
            expected = str(expected).strip() if expected else ""

            print(f"Running Row {i}: {input_text}")

            # 2. Input and Click Logic
            input_box.click()
            input_box.fill("") # Clear box before filling
            input_box.fill(input_text)
            
            translate_btn.click()

            # 3. Increased Wait Time
            page.wait_for_timeout(8000)

            actual = _read_output_with_wait(page, output_box, args.output_wait_ms)
            ws.cell(i, actual_col).value = actual

            if expected:
                status = "PASS" if actual == expected else "FAIL"
            else:
                status = "COLLECTED"

            ws.cell(i, status_col).value = status
            print(f" -> {status}")

            # Save after every row to prevent data loss
            wb.save(args.excel)

        print("Done!")

        if args.keep_open:
            print("Keeping browser open...")
            while True:
                page.wait_for_timeout(1000)

if __name__ == "__main__":
    run_test()